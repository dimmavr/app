import datetime
import tempfile
from io import BytesIO

from django.contrib.auth.models import User
from django.http import HttpResponse
from django.template.loader import render_to_string
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from rest_framework import filters, status, viewsets
from rest_framework.authtoken.models import Token
from rest_framework.authtoken.views import ObtainAuthToken
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.viewsets import ViewSet

from .models import Customer, Item, Order, OrderItem, Payment
from .serializers import (CustomerSerializer, ItemSerializer,
                          OrderItemCreateSerializer, OrderItemSerializer,
                          OrderSerializer, PaymentSerializer)


@permission_classes([AllowAny])
class CustomAuthToken(ObtainAuthToken):
    def post(self, request, *args, **kwargs):
        serializer = self.serializer_class(data=request.data,
                                           context={'request': request})
        serializer.is_valid(raise_exception=True)
        user = serializer.validated_data['user']
        token, created = Token.objects.get_or_create(user=user)
        return Response({'token': token.key})
    




class CustomerViewSet(viewsets.ModelViewSet):
    queryset = Customer.objects.all()
    serializer_class = CustomerSerializer
    filter_backends = [filters.SearchFilter, DjangoFilterBackend]
    search_fields = ['first_name', 'last_name', 'phone', 'tax_id']
    filterset_fields = ['first_name', 'last_name', 'phone', 'tax_id']

    @action(detail=True, methods=['get'])
    def debt(self, request, pk=None):
        try:
            customer = Customer.objects.get(pk=pk)
        except Customer.DoesNotExist:
            return Response({"error": "Ο πελάτης δεν βρέθηκε"}, status=404)

        orders = Order.objects.filter(customer=customer)
        total_orders = sum(order.total_amount() for order in orders)
        payments = Payment.objects.filter(order__customer=customer)
        total_paid = sum(payment.amount for payment in payments)
        remaining = total_orders - total_paid

        return Response({
            "customer": f"{customer.first_name} {customer.last_name}",
            "total_orders": total_orders,
            "total_paid": total_paid,
            "debt": remaining
        })
    

    @action(detail=False, methods=['get'])
    def export_pdf(self, request):
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)

        y = 800
        p.setFont("Helvetica-Bold", 14)
        p.drawString(100, y, "Λίστα Πελατών")
        y -= 40

        p.setFont("Helvetica", 10)
        customers = Customer.objects.all()

        for c in customers:
            if y < 50:
                p.showPage()
                y = 800
            p.drawString(50, y, f"ID: {c.id} | {c.first_name} {c.last_name}")
            y -= 15
            p.drawString(70, y, f"ΑΦΜ: {c.tax_id} | Τηλ: {c.phone} | Email: {c.email or '—'}")
            y -= 25

        p.save()
        buffer.seek(0)
        return HttpResponse(buffer, content_type='application/pdf')
    













class ItemViewSet(viewsets.ModelViewSet):
    queryset = Item.objects.all()
    serializer_class = ItemSerializer
    filter_backends = [filters.SearchFilter, DjangoFilterBackend]
    search_fields = ['name', 'description']
    filterset_fields = ['name']

    @action(detail=False, methods=['get'])
    def top_selling(self, request):
        stats = {}
        for oi in OrderItem.objects.all():
            name = oi.item.name
            stats[name] = stats.get(name, 0) + oi.quantity
        top = sorted(stats.items(), key=lambda x: x[1], reverse=True)[:5]
        return Response({k: v for k, v in top})

    @action(detail=False, methods=['get'])
    def sold_by_date(self, request):
        date_str = request.query_params.get('date')
        if not date_str:
            return Response({"error": "Δώσε ημερομηνία με ?date=YYYY-MM-DD"}, status=400)
        try:
            date_obj = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        except:
            return Response({"error": "Λάθος μορφή ημερομηνίας"}, status=400)

        items_today = OrderItem.objects.filter(order__date=date_obj)
        result = {}
        for item in items_today:
            name = item.item.name
            result[name] = result.get(name, 0) + item.quantity
        return Response(result)






class OrderViewSet(viewsets.ModelViewSet):
    queryset = Order.objects.all()
    serializer_class = OrderSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['customer', 'date']
    search_fields = ['customer__first_name', 'customer__last_name']
    

    @action(detail=False, methods=['get'])
    def today(self, request):
        today_date = datetime.date.today()
        todays_orders = Order.objects.filter(date=today_date)
        serializer = self.get_serializer(todays_orders, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def summary(self, request, pk=None):
        try:
            order = Order.objects.get(pk=pk)
        except Order.DoesNotExist:
            return Response({"error": "Η παραγγελία δεν βρέθηκε"}, status=404)

        items_data = [{
            "item": i.item.name,
            "quantity": i.quantity,
            "price": i.price
        } for i in order.items.all()]

        payments_data = [{
            "amount": p.amount,
            "date": p.date
        } for p in order.payments.all()]

        return Response({
            "order_id": order.id,
            "customer": f"{order.customer.first_name} {order.customer.last_name}",
            "date": order.date,
            "total_amount": order.total_amount(),
            "paid_amount": order.paid_amount(),
            "remaining_amount": order.remaining_amount(),
            "is_paid": order.is_paid(),
            "items": items_data,
            "payments": payments_data
        })


    @action(detail=False, methods=['get'])
    def export_pdf(self, request):
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)

        y = 800
        p.setFont("Helvetica-Bold", 14)
        p.drawString(100, y, "Λίστα Παραγγελιών")
        y -= 40

        p.setFont("Helvetica", 10)
        orders = Order.objects.all().select_related("customer")

        for order in orders:
            if y < 50:
                p.showPage()
                y = 800
            p.drawString(50, y, f"ID: {order.id} | Ημερομηνία: {order.date} | Πελάτης: {order.customer.first_name} {order.customer.last_name}")
            y -= 15
            p.drawString(70, y, f"Ποσό: {order.total_amount()} €  |  Πληρωμένο: {order.paid_amount()} €  | Υπόλοιπο: {order.remaining_amount()} €")
            y -= 25

        p.save()
        buffer.seek(0)
 
        return HttpResponse(buffer, content_type='application/pdf')
    


    @action(detail=False, methods=['get'])
    def by_date(self,request):
        date_str = request.query_params.get('date')
        if not date_str:
            return Response({"error": "Δώσε ημερομηνία με ?date=YYYY-MM-DD"}, status=400)
        try:
            date_obj = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return Response({"error": "Λάθος μορφή ημερομηνίας"}, status=400)   
        orders_today = Order.objects.filter(date=date_obj)
        serializer= self.get_serializer(orders_today, many=True)
        return Response(serializer.data)
    


    @action(detail=False, methods=['get'])
    def by_month(self, request):
        month = request.query_params.get('month')  # π.χ. ?month=2025-07
        if not month:
             return Response({"error": "Δώσε μήνα (YYYY-MM)"}, status=400)

        try:
            year, month = map(int, month.split('-'))
        except:
            return Response({"error": "Λάθος μορφή μήνα"}, status=400)

        orders = Order.objects.filter(date__year=year, date__month=month)
        serializer = self.get_serializer(orders, many=True)
        return Response(serializer.data)


    


    



class OrderItemViewSet(viewsets.ModelViewSet):
    queryset = OrderItem.objects.all()
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['order', 'item', 'order__date']

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return OrderItemCreateSerializer
        return OrderItemSerializer






class PaymentViewSet(viewsets.ModelViewSet):
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['date', 'amount', 'order']

    @action(detail=False, methods=['get'])
    def today(self, request):
        today = datetime.date.today()
        todays_payments = Payment.objects.filter(date=today)
        serializer = self.get_serializer(todays_payments, many=True)
        return Response(serializer.data)
    


    @action(detail=False, methods=['get'])
    def export_pdf(self, request):
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)

        y = 800
        p.setFont("Helvetica-Bold", 14)
        p.drawString(100, y, "Λίστα Πληρωμών")
        y -= 40

        p.setFont("Helvetica", 10)
        payments = Payment.objects.all().select_related("order")

        for pay in payments:
            if y < 50:
                p.showPage()
                y = 800
            p.drawString(50, y, f"ID: {pay.id} | Ημερομηνία: {pay.date}")
            y -= 15
            p.drawString(70, y, f"Ποσό: {pay.amount} € | Παραγγελία #{pay.order.id}")
            y -= 25

        p.save()
        buffer.seek(0)
        return HttpResponse(buffer, content_type='application/pdf')

    






    


class DashboardViewSet(viewsets.ViewSet):

    def list(self, request):
        return Response({
            "daily_sales": request.build_absolute_uri('daily_sales/'),
            "daily_payments": request.build_absolute_uri('daily_payments/'),
            "top_debtors": request.build_absolute_uri('top_debtors/')
        })

   

    @action(detail=False, methods=['get'])
    def daily_sales(self, request):
        today = datetime.date.today()
        orders = Order.objects.filter(date=today)
        total = sum(order.total_amount() for order in orders)
        return Response({
            "date": today,
            "total_sales": total,
            "total_orders": orders.count()
        })

    @action(detail=False, methods=['get'])
    def top_debtors(self, request):
        customers = Customer.objects.all()
        data = []
        for c in customers:
            total = sum(o.total_amount() for o in c.orders.all())
            paid = sum(o.paid_amount() for o in c.orders.all())
            if total - paid > 0:
                data.append({
                    "customer": f"{c.first_name} {c.last_name}",
                    "debt": total - paid
                })
        data.sort(key=lambda x: x["debt"], reverse=True)
        return Response(data[:5])

    @action(detail=False, methods=['get'])
    def daily_payments(self, request):
        today = datetime.date.today()
        payments = Payment.objects.filter(date=today)
        total = sum(p.amount for p in payments)
        return Response({
            "date": today,
            "total_payments": total,
            "count": payments.count()
        })

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        orders = Order.objects.all().select_related('customer')
    
        wb = Workbook()
        ws = wb.active
        ws.title = "Orders"

        ws.append([
        "ID", "Ημερομηνία", "Πελάτης", "Συνολικό Ποσό", "Πληρωμένο Ποσό", "Υπόλοιπο", "Εξοφλημένη"
        ])

        for order in orders:
            ws.append([
                order.id,
                str(order.date),
                f"{order.customer.first_name} {order.customer.last_name}",
                float(order.total_amount()),
                float(order.paid_amount()),
                float(order.remaining_amount()),
                "ΝΑΙ" if order.is_paid() else "ΟΧΙ"
            ])

        response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=orders.xlsx'
        wb.save(response)
        return response
    



    @action(detail=False, methods=['get'])
    def export_payments_excel(self, request):
        payments = Payment.objects.all().select_related('order')

        wb = Workbook()
        ws = wb.active
        ws.title = "Payments"

        ws.append(["ID", "Ποσό", "Ημερομηνία", "Παραγγελία"])

        for p in payments:
            ws.append([
                p.id,
                float(p.amount),
                str(p.date),
                f"#{p.order.id}"
        ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
        response['Content-Disposition'] = 'attachment; filename=payments.xlsx'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'])
    def export_summary(self, request):
        today = datetime.date.today()
        orders = Order.objects.filter(date=today)
        payments = Payment.objects.filter(date=today)
        customers = Customer.objects.all()

        total_sales = sum(order.total_amount() for order in orders)
        total_payments = sum(p.amount for p in payments)

        debtors = []
        for c in customers:
            total = sum(o.total_amount() for o in c.orders.all())
            paid = sum(o.paid_amount() for o in c.orders.all())
            if total - paid > 0:
                debtors.append((f"{c.first_name} {c.last_name}", total - paid))
        debtors.sort(key=lambda x: x[1], reverse=True)
        top_debtors = debtors[:5]

        wb = Workbook()
        ws = wb.active
        ws.title = "Σύνοψη"

        ws.append(["Ημερομηνία", str(today)])
        ws.append(["Σύνολο Πωλήσεων", total_sales])
        ws.append(["Σύνολο Πληρωμών", total_payments])
        ws.append([])
        ws.append(["Top Χρεωμένοι Πελάτες"])
        ws.append(["Πελάτης", "Χρέος"])
        for name, debt in top_debtors:
            ws.append([name, debt])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=summary.xlsx'
        wb.save(response)
        return response
    





    @action(detail=False, methods=['get'])
    def export_summary_pdf(self, request):
        today = datetime.date.today()
        orders = Order.objects.filter(date=today)
        payments = Payment.objects.filter(date=today)
        customers = Customer.objects.all()

        total_sales = sum(order.total_amount() for order in orders)
        total_payments = sum(p.amount for p in payments)

        # Υπολογισμός top χρεωμένων
        debtors = []
        for c in customers:
            total = sum(o.total_amount() for o in c.orders.all())
            paid = sum(o.paid_amount() for o in c.orders.all())
            debt = total - paid
            if debt > 0:
                debtors.append((f"{c.first_name} {c.last_name}", debt))
                debtors.sort(key=lambda x: x[1], reverse=True)
                top_debtors = debtors[:5]

            # Δημιουργία PDF
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        y = 800

        p.setFont("Helvetica-Bold", 14)
        p.drawString(100, y, "Ημερήσια Σύνοψη")
        y -= 40

        p.setFont("Helvetica", 11)
        p.drawString(50, y, f"Ημερομηνία: {today}")
        y -= 20
        p.drawString(50, y, f"Σύνολο Πωλήσεων: {total_sales} €")
        y -= 20
        p.drawString(50, y, f"Σύνολο Πληρωμών: {total_payments} €")
        y -= 40

        p.setFont("Helvetica-Bold", 12)
        p.drawString(50, y, "Top Χρεωμένοι Πελάτες:")
        y -= 25

        p.setFont("Helvetica", 10)
        for name, debt in top_debtors:
            if y < 50:
                p.showPage()
                y = 800
            p.drawString(70, y, f"{name} - Χρέος: {debt} €")
            y -= 20

        p.save()
        buffer.seek(0)

        return HttpResponse(buffer, content_type='application/pdf')
    


    @action(detail=False, methods=['get'])
    def overdue_debtors(self, request):
        days = int(request.query_params.get('days', 30))
        cutoff = datetime.date.today() - datetime.timedelta(days=days)

        overdue_customers = []
        for customer in Customer.objects.all():
            old_orders = customer.orders.filter(date__lte=cutoff)
            total = sum(o.total_amount() for o in old_orders)
            paid = sum(o.paid_amount() for o in old_orders)
            debt = total - paid
            if debt > 0:
                overdue_customers.append({
                  "customer": f"{customer.first_name} {customer.last_name}",
                  "debt": debt,
                  "orders_count": old_orders.count()
               })

        return Response(overdue_customers)



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def change_password(request):
    user = request.user
    old_password = request.data.get('old_password')
    new_password = request.data.get('new_password')

    if not user.check_password(old_password):
        return Response({"error": "Λάθος τρέχων κωδικός"}, status=status.HTTP_400_BAD_REQUEST)

    user.set_password(new_password)
    user.save()
    return Response({"success": "Ο κωδικός άλλαξε επιτυχώς"})



@api_view(['POST'])
@permission_classes([AllowAny])
def register_user(request):
    username = request.data.get('username')
    password = request.data.get('password')
    email = request.data.get('email')
    first_name = request.data.get('first_name', '')
    last_name = request.data.get('last_name', '')

    if not username or not password or not email:
        return Response({"error": "username, password και email είναι υποχρεωτικά"}, status=400)

    if User.objects.filter(username=username).exists():
        return Response({"error": "Το όνομα χρήστη υπάρχει ήδη"}, status=400)

    if User.objects.filter(email=email).exists():
        return Response({"error": "Αυτό το email χρησιμοποιείται ήδη"}, status=400)

    user = User.objects.create_user(
        username=username,
        password=password,
        email=email,
        first_name=first_name,
        last_name=last_name
    )

    return Response({
        "success": f"Ο χρήστης {user.username} δημιουργήθηκε επιτυχώς",
        "username": user.username,
        "email": user.email,
        "first_name": user.first_name,
        "last_name": user.last_name
    }, status=201)



@action(detail=False, methods=['get'])
def debtors_all(self, request):
    customers = Customer.objects.all()
    data = []

    for c in customers:
        total = sum(o.total_amount() for o in c.orders.all())
        paid = sum(o.paid_amount() for o in c.orders.all())
        debt = total - paid

        if debt > 0:
            data.append({
                "customer": f"{c.first_name} {c.last_name}",
                "debt": round(debt, 2)
            })

    data.sort(key=lambda x: x["debt"], reverse=True)
    return Response(data)






@action(detail=False, methods=['get'])
def sales_report(self, request):
    date_str = request.query_params.get('date')
    month_str = request.query_params.get('month')
    customer_id = request.query_params.get('customer')

    orders = Order.objects.all()

    if date_str:
        try:
            date = datetime.strptime(date_str, "%Y-%m-%d").date()
            orders = orders.filter(date=date)
        except ValueError:
            return Response({"error": "Invalid date format. Use YYYY-MM-DD."}, status=400)

    elif month_str:
        try:
            month = datetime.strptime(month_str, "%Y-%m").date()
            orders = orders.filter(date__year=month.year, date__month=month.month)
        except ValueError:
            return Response({"error": "Invalid month format. Use YYYY-MM."}, status=400)

    if customer_id:
        orders = orders.filter(customer_id=customer_id)

    total_sales = 0
    total_orders = orders.count()
    items_summary = {}

    for order in orders:
        total_sales += float(order.total_amount)
        for order_item in order.items.all():
            name = order_item.item.name
            qty = order_item.quantity
            total = order_item.total_price()

            if name not in items_summary:
                items_summary[name] = {"quantity": 0, "total": 0}
            items_summary[name]["quantity"] += qty
            items_summary[name]["total"] += float(total)

    items_list = [
        {"name": name, "quantity": data["quantity"], "total": round(data["total"], 2)}
        for name, data in items_summary.items()
    ]

    return Response({
        "total_sales": round(total_sales, 2),
        "total_orders": total_orders,
        "items": items_list
    })